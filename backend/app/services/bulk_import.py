import csv
import io
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.employee import Employee, RoleEnum
from app.models.project import EmployeeProject, Project
from app.utils.auth import hash_password

HEADER_ALIASES = {
    "project name": "project_name",
    "project number": "project_number",
    "user name": "user_name",
    "mobile number": "mobile_number",
    "password": "password",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_mobile(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        key = _clean(header).lower()
        if key in HEADER_ALIASES:
            mapping[idx] = HEADER_ALIASES[key]
    return mapping


def _row_from_values(values: list[Any], header_map: dict[int, str]) -> dict[str, str]:
    row: dict[str, str] = {}
    for idx, field in header_map.items():
        if idx < len(values):
            row[field] = _clean(values[idx])
    return row


def parse_upload_file(filename: str, content: bytes) -> list[dict[str, str]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _parse_csv(content)
    if lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return _parse_xlsx(content)
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("The file is empty.")

    header_map = _map_headers(rows[0])
    if len(header_map) < 5:
        raise ValueError(
            "Missing required columns. Expected: Project Name, Project Number, "
            "User Name, Mobile Number, Password."
        )

    parsed: list[dict[str, str]] = []
    for values in rows[1:]:
        row = _row_from_values(values, header_map)
        if any(row.values()):
            parsed.append(row)
    return parsed


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    sheet_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not sheet_rows:
        raise ValueError("The file is empty.")

    header_map = _map_headers(list(sheet_rows[0]))
    if len(header_map) < 5:
        raise ValueError(
            "Missing required columns. Expected: Project Name, Project Number, "
            "User Name, Mobile Number, Password."
        )

    parsed: list[dict[str, str]] = []
    for values in sheet_rows[1:]:
        row = _row_from_values(list(values or []), header_map)
        if any(row.values()):
            parsed.append(row)
    return parsed


def import_employee_rows(db: Session, rows: list[dict[str, str]]) -> dict[str, Any]:
    projects_created = 0
    employees_created = 0
    assignments_created = 0
    rows_processed = 0
    rows_skipped = 0
    errors: list[str] = []

    project_cache: dict[str, Project] = {}
    employee_cache: dict[str, Employee] = {}

    for line_no, row in enumerate(rows, start=2):
        project_name = row.get("project_name", "").strip()
        project_number = row.get("project_number", "").strip()
        user_name = row.get("user_name", "").strip()
        mobile_number = _normalize_mobile(row.get("mobile_number", ""))
        password = row.get("password", "").strip()

        if not project_name and not project_number and not user_name and not mobile_number:
            rows_skipped += 1
            continue

        if not project_number or not project_name:
            errors.append(f"Row {line_no}: project name and project number are required.")
            rows_skipped += 1
            continue
        if not user_name or not mobile_number:
            errors.append(f"Row {line_no}: user name and mobile number are required.")
            rows_skipped += 1
            continue
        if not re.fullmatch(r"\d{10}", mobile_number):
            errors.append(f"Row {line_no}: mobile number must be 10 digits ({mobile_number}).")
            rows_skipped += 1
            continue
        if len(password) < 6:
            errors.append(f"Row {line_no}: password must be at least 6 characters.")
            rows_skipped += 1
            continue

        try:
            with db.begin_nested():
                project = project_cache.get(project_number)
                if not project:
                    project = (
                        db.query(Project)
                        .filter(Project.project_number == project_number)
                        .first()
                    )
                    if not project:
                        project = Project(
                            project_number=project_number,
                            project_name=project_name,
                        )
                        db.add(project)
                        db.flush()
                        projects_created += 1
                    project_cache[project_number] = project

                employee = employee_cache.get(mobile_number)
                if not employee:
                    employee = (
                        db.query(Employee)
                        .filter(Employee.mobile_number == mobile_number)
                        .first()
                    )
                    if not employee:
                        employee = Employee(
                            name=user_name,
                            mobile_number=mobile_number,
                            password_hash=hash_password(password),
                            role=RoleEnum.employee,
                        )
                        db.add(employee)
                        db.flush()
                        employees_created += 1
                    elif employee.name != user_name:
                        employee.name = user_name
                    employee_cache[mobile_number] = employee

                existing_assignment = (
                    db.query(EmployeeProject)
                    .filter(
                        EmployeeProject.employee_id == employee.id,
                        EmployeeProject.project_id == project.id,
                    )
                    .first()
                )
                if not existing_assignment:
                    db.add(
                        EmployeeProject(
                            employee_id=employee.id,
                            project_id=project.id,
                        )
                    )
                    assignments_created += 1

            rows_processed += 1
        except Exception as exc:
            errors.append(f"Row {line_no}: {exc}")
            rows_skipped += 1

    db.commit()
    return {
        "projects_created": projects_created,
        "employees_created": employees_created,
        "assignments_created": assignments_created,
        "rows_processed": rows_processed,
        "rows_skipped": rows_skipped,
        "errors": errors[:20],
    }
